from __future__ import print_function
from openpyxl import load_workbook, Workbook
from copy import copy
import os


def get_style(cell):
    # WARNING:
    # FROM: https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
    # cell.font or cell.border is an instance of StyleProxy
    # if save workbook with that type, you will get an exception. 
    # You must copy it to new cell, like this: new_cell.font = copy(cell.font)
    return {
        'font'          : copy(cell.font),
        'border'        : copy(cell.border),
        'fill'          : copy(cell.fill),
        'number_format' : copy(cell.number_format),
        'protection'    : copy(cell.protection),
        'alignment'     : copy(cell.alignment)
    }

def apply_style(cell, style_obj):
    cell.font          = style_obj['font']
    cell.border        = style_obj['border']
    cell.fill          = style_obj['fill']
    cell.number_format = style_obj['number_format']
    cell.protection    = style_obj['protection']
    cell.alignment     = style_obj['alignment']
    return cell

class ExcelInterface():

    def __init__(self, path, sheet, START_ROW=0):

        # If there is existing worksheet load it
        # If there is no worksheet create it
        if os.path.isfile(path):
            self.wb = load_workbook(path)
        else:
            print('[WARNING] file at %s not found, create new file' % path)
            self.wb = Workbook()
            self.wb.create_sheet(title=sheet)
        self.ws = self.wb[sheet]
        self.START_ROW = START_ROW
        self.path = path

        self.ROW_TEMPLATE = []
        self.ROW_STYLE = []
        return


    # Excel interface to read as string
    def read_header(self, char_idx_start, char_idx_stop):
        # args as character range
        output = []
        # data structure when read range is ((c1r1, c2r1,...), (c1r2, c2r2,...))
        for cell in self.ws[char_idx_start: char_idx_stop][0]:
            output.append(str(cell.value).strip().lower())
        return output

    def read_data(self, START_ROW=0):
        # output as list of list
        self.START_ROW = START_ROW
        output = []
        for row in self.ws.iter_rows(min_row = START_ROW):
            row_value = [ str(r.value).strip() for r in row ]
            output.append(row_value)
        return output

    def read_dict(self, START_ROW=0):
        self.START_ROW = START_ROW
        output = []
        headers = [str(x.value).strip() for x in self.ws[START_ROW]]
        for row in self.ws.iter_rows(min_row = START_ROW + 1):
            row_value = [ str(r.value).strip() for r in row ]
            output.append(dict(zip(headers, row_value)))
        return output

    def write_dict(self, dicts):
        # TODO check consistency
        # TODO deal with default Sheet work sheet
        headers = dicts[0].keys()
        for j, header in enumerate(headers):
            col = j+1
            self.ws.cell(column=col, row=1, value=header)
        for i, dct in enumerate(dicts):
            for j, header in enumerate(headers):
                # Excel count from 1 and skip header
                row = i+2
                col = j+1
                _ = self.ws.cell(column=col, row=row, value=dct[header])
        self.save(filename=self.path)

    def format(self, contexts, ENVIRONMENT):
        for cell in self.ws[self.START_ROW]:
            self.ROW_TEMPLATE.append(str(cell.value))
            self.ROW_STYLE.append(get_style(cell))
        for i, context in enumerate(contexts):
            for j in range(0, len(self.ROW_TEMPLATE)):
                # print(ROW_TEMPLATE[j].format(context))
                context['k'] = i+1
                # WARNING: Excel count 1,2,3,...
                _cell = self.ws.cell(row=self.START_ROW+i,column=j+1)
                _cell.value = self.ROW_TEMPLATE[j].format(**context)
                self.ws.row_dimensions[self.START_ROW+i+1].height = self.ws.row_dimensions[self.START_ROW].height
                apply_style(_cell, self.ROW_STYLE[j])
                #ws.cell(row=i,column=j,value=ROW_TEMPLATE[j].format(context)
        # hard code for SIT/UAT doc
        self.ws['A1'] = self.ws['A1'].value.format(ENVIRONMENT=ENVIRONMENT.upper().replace('DL',''))
        return self

    def save(self, filename):
        self.wb.save(filename)
        return self


if __name__ == '__main__':
    TESTDIR='./'
    test_dicts = [{'a':1,'b':2,'c':3},{'a':4, 'b':5, 'c':6},{'a':7,'b':8,'c':9}]
    e1 = ExcelInterface(path='test_excel_01.xlsx', sheet='test').write_dict(test_dicts)
    e2 = ExcelInterface(path='test_excel_01.xlsx', sheet='test')
    print(e2.read_dict(START_ROW=1))
    print(test_dicts)
    assert(e1==e2.read_dict(START_ROW=2))
    assert(test_dicts.keys() == e2.read_header(START_ROW=1))

    print('clean up test dir')

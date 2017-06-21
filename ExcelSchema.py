# TEMP COMMIT TO CONTINUE WORK AT OFFICE

from openpyxl import load_workbook

class ExcelSchema():
        def __init__(self, spec):
                ''' [
                {'column_name': 'col1', 'data_type': 'integer'}
                {'column_name': 'col2', 'data_type': 'number'}
                {'column_name': 'col3', 'data_type': 'string'}
                {'column_name': 'col3', 'data_type': 'decimal'}
                ] '''
                return self

        def load(self, filename):
                wb = load_workbook(filename, read_only=True)
                self._validation_hook()

        def _validation_hook():


        def pre_processor(self):
                return self

        def post_processor(self):
                return self

if __name__ == '__main__':
        wb2 = load_workbook('test.xlsx', read_only=True)
        print wb2.get_sheet_names()
        ws = wb2[wb2.get_sheet_names()[0]]
        for row in ws.rows:
                for cell in row:
                        print cell.value
